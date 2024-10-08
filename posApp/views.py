from pickle import FALSE
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse, JsonResponse
from flask import jsonify
from posApp.models import Category, Mensualidad, PlanInscripcion, Products, Sales, Salida, salesItems, Clientes, Genero, Levels, FormaPago, Sucursal, UserProfile
from django.db.models import Count, Sum, Q
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import redirect
import json, sys
from datetime import date, datetime
from django.contrib.auth.models import User
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from posApp.models import Sucursal, UserProfile
from .forms import MensualidadForm, PlanInscripcionForm, SalidaForm, SellerUserCreationForm, SucursalForm
from django.views.generic import CreateView
from django.urls import reverse,  reverse_lazy
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from datetime import datetime
from calendar import monthrange
from django.contrib.auth.views import LoginView

# Login
def login_user(request):
    # Log out any existing user
    logout(request)
    
    resp = {"status": 'failed', 'msg': ''}
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        
        if user is not None:
            if user.is_active:
                login(request, user)
                resp['status'] = 'success'
                
                try:
                    # Retrieve the user's profile
                    user_profile = UserProfile.objects.get(user=user)
                    
                    # Check if the user is a manager using the user profile
                    if user_profile.is_manager:
                        resp['redirect_url'] = reverse('cambiar-sucursal')  # Redirect to change branch for managers
                    else:
                        resp['redirect_url'] = reverse('home-page')  # Redirect to home page for regular users
                except UserProfile.DoesNotExist:
                    # Handle case where user does not have a profile associated (fallback behavior)
                    resp['msg'] = "User profile not found. Please contact support."
            else:
                resp['msg'] = "Your account is inactive."
        else:
            resp['msg'] = "Incorrect username or password."
    
    return HttpResponse(json.dumps(resp), content_type='application/json')

class CustomLoginView(LoginView):
    template_name = 'posApp/login.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse_lazy('home-page')

def logoutuser(request):
    logout(request)
    return redirect('/')


# Vista para manejar el error 404
def error_404_view(request, exception):
    return render(request, 'posApp/404.html', {}, status=404)

# Vista para manejar el error 500
def error_500_view(request):
    return render(request, 'posApp/404.html', {}, status=500)


@login_required
def cambiar_sucursal(request):
    # Accede al perfil del usuario
    try:
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return redirect('home-page')  # Si no tiene un perfil, redirigir a la página de inicio

    # Verifica si el usuario es un gerente
    if not user_profile.is_manager:
        return redirect('home-page')  # Redirigir si no es un gerente

    # Procesar el formulario si se envía una sucursal
    if request.method == 'POST':
        sucursal_id = request.POST.get('sucursal')
        try:
            sucursal = Sucursal.objects.get(id=sucursal_id)
            # Guarda la sucursal en la sesión
            request.session['sucursal_id'] = sucursal.id
            request.session['sucursal_nombre'] = sucursal.nombre
            return redirect('home-page')
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no encontrada.'})

    # Si no es POST, mostrar las sucursales disponibles para cambiar
    sucursales = Sucursal.objects.all()
    return render(request, 'posApp/cambiar_sucursal.html', {'sucursales': sucursales})



class SellerUserCreateView(CreateView):
    model = User
    form_class = SellerUserCreationForm
    template_name = 'posApp/create_seller_user.html'
    success_url = reverse_lazy('home-page')  # Cambia esto a donde quieras redirigir después de la creación

    def form_valid(self, form):
        # Guardamos primero el usuario
        response = super().form_valid(form)

        # Crear el perfil del usuario y asignar los valores adicionales
        user = self.object  # Esto es el usuario que acabamos de crear
        sucursal = form.cleaned_data.get('sucursal')  # Obtener la sucursal seleccionada del formulario

        # Crear el perfil del usuario con is_seller=True
        UserProfile.objects.create(user=user, sucursal=sucursal, is_seller=True)

        return response

@login_required
def home(request):
    user = request.user

    # Verifica si el usuario tiene un perfil asociado
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asociada, a menos que sea un gerente
    if not user_profile.is_manager and not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determina la sucursal a usar
    if user_profile.is_manager:
        # Obtiene la sucursal desde la sesión
        sucursal_id = request.session.get('sucursal_id')
        if not sucursal_id:
            return render(request, 'posApp/error.html', {'message': 'No se ha seleccionado una sucursal.'})

        try:
            sucursal_a_usar = Sucursal.objects.get(id=sucursal_id)
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no encontrada en la sesión.'})
    else:
        # Para usuarios no gerentes, usa la sucursal del perfil
        sucursal_a_usar = user_profile.sucursal

    # Obtener la zona horaria de Ciudad de México
    tz = timezone.get_fixed_timezone(-6 * 60)  # UTC-6 es la zona horaria de Ciudad de México

    # Obtener la fecha y hora actual en la zona horaria de Ciudad de México
    now = timezone.now().astimezone(tz)
    current_year = now.year
    current_month = now.month
    current_day = now.day

    # Filtrar categorías, productos y ventas por la sucursal del perfil del usuario
    categories_count = Category.objects.filter(sucursal=sucursal_a_usar).count()
    products_count = Products.objects.filter(sucursal=sucursal_a_usar).count()

    sales_filters = {
        'sucursal': sucursal_a_usar,
        'date_added__year': current_year,
        'date_added__month': current_month,
    }

    today_sales = Sales.objects.filter(**sales_filters, date_added__day=current_day)
    month_sales = Sales.objects.filter(**sales_filters)
    month_sales_cash = month_sales.filter(tipoPago=1)
    month_sales_bank = month_sales.filter(tipoPago=2)

    total_sales = today_sales.aggregate(total=Sum('grand_total'))['total'] or 0
    month_sales_total = month_sales.aggregate(total=Sum('grand_total'))['total'] or 0
    month_sales_cash_total = month_sales_cash.aggregate(total=Sum('grand_total'))['total'] or 0
    month_sales_bank_total = month_sales_bank.aggregate(total=Sum('grand_total'))['total'] or 0

    user_info = {
        'username': user.username,
        'email': user.email,
        'is_staff': user.is_staff,
        'is_superuser': user.is_superuser,
        'sucursal': sucursal_a_usar.nombre if sucursal_a_usar else 'No asignada'
    }

    context = {
        'page_title': 'Home',
        'categories': categories_count,
        'products': products_count,
        'transaction': today_sales.count(),
        'total_sales': total_sales,
        'month_sales': month_sales_total,
        'month_sales_cash': month_sales_cash_total,
        'month_sales_bank': month_sales_bank_total,
        'user_info': user_info,
        'sucursal': sucursal_a_usar,
    }

    return render(request, 'posApp/home.html', context)


def about(request):
    context = {
        'page_title':'About',
    }
    return render(request, 'posApp/about.html',context)

#Categories
@login_required
def category(request):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})
    
    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})
    
    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal  # Si necesitas manejo de gerente, deberías hacerlo aquí

    # Filtrar las categorías por la sucursal del perfil del usuario
    category_list = Category.objects.filter(sucursal=sucursal_a_usar)

    context = {
        'page_title': 'Categorías',
        'category': category_list,
    }

    return render(request, 'posApp/category.html', context)

@login_required
def manage_category(request):
    category = None

    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal

    # Si el usuario es un gerente y la sucursal está en la sesión, se usa esa sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    if request.method == 'GET':
        data = request.GET
        id = data.get('id', '')
        if id.isnumeric():
            # Verificar que la categoría pertenezca a la sucursal del perfil del usuario
            category = Category.objects.filter(id=id, sucursal=sucursal_a_usar).first()
            if not category:
                return render(request, 'posApp/error.html', {
                    'message': 'Categoría no encontrada o no tienes permiso para acceder a ella.'
                })

    context = {
        'category': category
    }
    return render(request, 'posApp/manage_category.html', context)

@login_required
def save_category(request):
    data = request.POST
    resp = {'status': 'failed'}

    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        resp['msg'] = 'Error: el usuario no tiene un perfil asignado.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        resp['msg'] = 'Error: el perfil no tiene una sucursal asignada.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            resp['msg'] = 'Sucursal no válida en la sesión.'
            return HttpResponse(json.dumps(resp), content_type="application/json")

    try:
        category_id = data.get('id', '')
        name = data.get('name', '')
        description = data.get('description', '')
        status = data.get('status', '')

        if category_id.isnumeric() and int(category_id) > 0:
            # Intentar actualizar la categoría existente dentro de la sucursal del usuario
            category = Category.objects.filter(id=category_id, sucursal=sucursal_a_usar).first()
            if category:
                category.name = name
                category.description = description
                category.status = status
                category.save()
                resp['status'] = 'success'
                messages.success(request, 'Categoría actualizada con éxito.')
            else:
                resp['msg'] = 'Categoría no encontrada o no pertenece a su sucursal.'
        else:
            # Crear una nueva categoría asignándola a la sucursal del usuario
            new_category = Category(
                name=name,
                description=description,
                status=status,
                sucursal=sucursal_a_usar
            )
            new_category.save()
            resp['status'] = 'success'
            messages.success(request, 'Categoría creada con éxito.')

    except Exception as e:
        resp['status'] = 'failed'
        resp['msg'] = str(e)

    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def delete_category(request):
    data = request.POST
    resp = {'status': ''}

    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        resp['status'] = 'failed'
        resp['msg'] = 'Error: el usuario no tiene un perfil asignado.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        resp['status'] = 'failed'
        resp['msg'] = 'Error: el perfil no tiene una sucursal asignada.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            resp['status'] = 'failed'
            resp['msg'] = 'Sucursal no válida en la sesión.'
            return HttpResponse(json.dumps(resp), content_type="application/json")

    category_id = data.get('id', '')
    if not category_id.isnumeric():
        resp['status'] = 'failed'
        resp['msg'] = 'ID inválido.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    try:
        # Solo permite eliminar la categoría si pertenece a la sucursal del usuario
        category = Category.objects.filter(id=category_id, sucursal=sucursal_a_usar).first()
        if category:
            category.delete()
            resp['status'] = 'success'
            messages.success(request, 'Categoría eliminada con éxito.')
        else:
            resp['status'] = 'failed'
            resp['msg'] = 'Categoría no encontrada o no pertenece a su sucursal.'
    except Exception as e:
        resp['status'] = 'failed'
        resp['msg'] = str(e)

    return HttpResponse(json.dumps(resp), content_type="application/json")


@login_required
def delete_cliente(request):
    data = request.POST
    resp = {'status': ''}

    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        resp['status'] = 'failed'
        resp['msg'] = 'Error: el usuario no tiene un perfil asignado.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        resp['status'] = 'failed'
        resp['msg'] = 'Error: el perfil no tiene una sucursal asignada.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            resp['status'] = 'failed'
            resp['msg'] = 'Sucursal no válida en la sesión.'
            return HttpResponse(json.dumps(resp), content_type="application/json")

    cliente_id = data.get('id', '')
    if not cliente_id.isnumeric():
        resp['status'] = 'failed'
        resp['msg'] = 'ID inválido.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    try:
        # Solo permite eliminar el cliente si pertenece a la sucursal del usuario
        cliente = Clientes.objects.filter(id=cliente_id, sucursal=sucursal_a_usar).first()
        if cliente:
            cliente.delete()
            resp['status'] = 'success'
            messages.success(request, 'Cliente eliminado con éxito.')
        else:
            resp['status'] = 'failed'
            resp['msg'] = 'Cliente no encontrado o no pertenece a su sucursal.'
    except Exception as e:
        resp['status'] = 'failed'
        resp['msg'] = str(e)

    return HttpResponse(json.dumps(resp), content_type="application/json")


# Products
@login_required
def products(request):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    # Obtener los productos de la sucursal determinada
    product_list = Products.objects.filter(sucursal=sucursal_a_usar)

    context = {
        'page_title': 'Productos',
        'products': product_list,
    }

    return render(request, 'posApp/products.html', context)

@login_required
def plans(request):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    # Obtener los planes de la sucursal determinada
    plan_list = PlanInscripcion.objects.filter(sucursal=sucursal_a_usar)

    context = {
        'page_title': 'Planes de Inscripción',
        'plans': plan_list,
    }

    return render(request, 'posApp/plans.html', context)

def editar_plan(request, pk):
    plan = get_object_or_404(PlanInscripcion, pk=pk)
    if request.method == 'POST':
        form = PlanInscripcionForm(request.POST, instance=plan)
        if form.is_valid():
            form.save()
            return redirect('plans-page')  # Redirige a la lista de planes después de la edición
    else:
        form = PlanInscripcionForm(instance=plan)
    
    return render(request, 'posApp/editar_plan.html', {'form': form, 'plan': plan})

@csrf_exempt
def delete_plan(request):
    if request.method == "POST":
        try:
            # Intentamos obtener el perfil del usuario
            user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
            return JsonResponse({'status': 'failed', 'msg': 'Error: el usuario no tiene un perfil asignado.'})

        # Asegúrate de que el perfil tenga una sucursal asignada
        if not user_profile.sucursal:
            return JsonResponse({'status': 'failed', 'msg': 'Error: el perfil no tiene una sucursal asignada.'})

        # Determinar la sucursal a usar
        sucursal_a_usar = user_profile.sucursal
        if user_profile.is_manager and 'sucursal_id' in request.session:
            try:
                sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
            except Sucursal.DoesNotExist:
                return JsonResponse({'status': 'failed', 'msg': 'Sucursal no válida en la sesión.'})

        id = request.POST.get('id')
        try:
            # Verificar que el plan de inscripción pertenezca a la sucursal del usuario
            plan = PlanInscripcion.objects.get(id=id, sucursal=sucursal_a_usar)
            plan.delete()
            return JsonResponse({'status': 'success'})
        except PlanInscripcion.DoesNotExist:
            return JsonResponse({'status': 'failed', 'msg': 'El plan de inscripción no existe o no pertenece a su sucursal.'})
    
    return JsonResponse({'status': 'failed', 'msg': 'Método no permitido.'})


@login_required
def clients(request):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    # Filtrar clientes por la sucursal del perfil del usuario
    all_clients = Clientes.objects.filter(sucursal=sucursal_a_usar)

    # Aplicar filtros adicionales si son proporcionados
    estado_filtro = request.GET.get('estado_mensualidad')
    horario_filtro = request.GET.get('horario')
    plan_inscripcion_filtro = request.GET.get('plan_inscripcion')
    query = request.GET.get('q')

    if estado_filtro:
        all_clients = all_clients.filter(estado_mensualidad=estado_filtro)
    if horario_filtro:
        all_clients = all_clients.filter(horario__icontains=horario_filtro)
    if plan_inscripcion_filtro:
        all_clients = all_clients.filter(plan_inscripcion__id=plan_inscripcion_filtro)

    if query:
        query = query.lower()
        search_terms = query.split()  # Dividir la consulta en palabras
        all_clients = [
            client for client in all_clients 
            if any(term in client.nombre.lower() or term in client.apellido_paterno.lower() or term in client.apellido_materno.lower() 
                for term in search_terms)
        ]

    horarios = [ 
        "AGUAS ABIERTAS", "EQUIPO A", "EQUIPO B", "PRESELECTIVOS", 
        "7:00 PRINCIPIANTES", "8:00 PRINCIPIANTES", "9:00 PRINCIPIANTES", 
        "7:00 INTERMEDIOS", "8:00 INTERMEDIOS", "9:00 INTERMEDIOS", 
        "7:00 AVANZADOS", "8:00 AVANZADOS", "9:00 AVANZADOS", 
        "7:00", "08:00", "09:00", "10:00", "11:00", "12:00", 
        "13:00", "14:00", "15:00", "16:00", "17:00", "18:00", 
        "19:00", "20:00", "21:00", 
        "19:00 PRINCIPIANTES", "20:00 PRINCIPIANTES", "21:00 PRINCIPIANTES", 
        "19:00 INTERMEDIOS", "20:00 INTERMEDIOS", "21:00 INTERMEDIOS", 
        "19:00 AVANZADOS", "20:00 AVANZADOS", "21:00 AVANZADOS" 
    ]

    planes_inscripcion = PlanInscripcion.objects.filter(sucursal=sucursal_a_usar)

    context = {
        'page_title': 'Clientes',
        'clients': all_clients,
        'query': query,
        'estado_filtro': estado_filtro,
        'horario_filtro': horario_filtro,
        'plan_inscripcion_filtro': plan_inscripcion_filtro,
        'horarios': horarios,
        'planes_inscripcion': planes_inscripcion,
    }

    return render(request, 'posApp/clients.html', context)


@login_required
def add_payment(request, cliente_id):
    client = Clientes.objects.get(pk=cliente_id)
    if request.method == 'POST':
        form = MensualidadForm(request.POST)
        if form.is_valid():
            mensualidad = form.save(commit=False)
            mensualidad.cliente = client
            mensualidad.fecha_pago = timezone.now()
            mensualidad.pagado = True
            mensualidad.save()
            return redirect('cliente_mensualidades', cliente_id=cliente_id)
    else:
        form = MensualidadForm()
    
    context = {
        'client': client,
        'form': form,
    }
    return render(request, 'posApp/add_payment.html', context)

@login_required
def manage_clients(request):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    client = None
    genero = Genero.objects.all()
    nivel = Levels.objects.all()
    planes_de_inscripccion = PlanInscripcion.objects.filter(sucursal=sucursal_a_usar)

    if request.method == 'GET':
        data = request.GET
        id = data.get('id', '')
        if id.isnumeric():
            # Asegurarse de que el cliente pertenece a la sucursal del perfil del usuario
            client = Clientes.objects.filter(id=id, sucursal=sucursal_a_usar).first()
            if not client:
                return render(request, 'posApp/error.html', {
                    'message': 'Cliente no encontrado o no pertenece a su sucursal.'
                })
                
    horarios = [
        "AGUAS ABIERTAS", "EQUIPO A", "EQUIPO B", "PRESELECTIVOS", 
        "7:00 PRINCIPIANTES", "8:00 PRINCIPIANTES", "9:00 PRINCIPIANTES", 
        "7:00 INTERMEDIOS", "8:00 INTERMEDIOS", "9:00 INTERMEDIOS", 
        "7:00 AVANZADOS", "8:00 AVANZADOS", "9:00 AVANZADOS", 
        "7:00", "08:00", "09:00", "10:00", "11:00", "12:00", 
        "13:00", "14:00", "15:00", "16:00", "17:00", "18:00", 
        "19:00", "20:00", "21:00", 
        "19:00 PRINCIPIANTES", "20:00 PRINCIPIANTES", "21:00 PRINCIPIANTES", 
        "19:00 INTERMEDIOS", "20:00 INTERMEDIOS", "21:00 INTERMEDIOS", 
        "19:00 AVANZADOS", "20:00 AVANZADOS", "21:00 AVANZADOS"
    ]
    
    context = {
        'client': client,
        'generos': genero,
        'niveles': nivel,
        'planes': planes_de_inscripccion,
        'horarios': horarios,
    }
    return render(request, 'posApp/manage_client.html', context)

@login_required
def save_client(request):
    data = request.POST
    resp = {'status': 'failed'}
    id = data.get('id', '')

    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        resp['msg'] = 'Error: el usuario no tiene un perfil asignado.'
        return JsonResponse(resp)

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        resp['msg'] = 'Error: el perfil no tiene una sucursal asignada.'
        return JsonResponse(resp)

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            resp['msg'] = 'Sucursal no válida en la sesión.'
            return JsonResponse(resp)

    # Verificar si el cliente ya existe en la misma sucursal
    if id.isnumeric() and int(id) > 0:
        check = Clientes.objects.exclude(id=id).filter(
            nombre=data['nombre'], 
            apellido_paterno=data['apellido_paterno'], 
            apellido_materno=data['apellido_materno'],
            sucursal=sucursal_a_usar
        ).all()
    else:
        check = Clientes.objects.filter(
            nombre=data['nombre'], 
            apellido_paterno=data['apellido_paterno'], 
            apellido_materno=data['apellido_materno'],
            sucursal=sucursal_a_usar
        ).all()

    if len(check) > 0:
        resp['msg'] = "El cliente ya existe, prueba cambiando los datos."
    else:
        genero = Genero.objects.filter(id=data['genero_id']).first()
        plan_inscripcion = PlanInscripcion.objects.filter(id=data['plan_inscripcion']).first()
        try:
            if id.isnumeric() and int(id) > 0:
                # Actualizar cliente existente
                Clientes.objects.filter(id=id).update(
                    nombre=data['nombre'], 
                    genero=genero, 
                    apellido_paterno=data['apellido_paterno'], 
                    apellido_materno=data['apellido_materno'], 
                    celular=data['celular'], 
                    fecha_nacimiento=data['fecha_nacimiento'],
                    codigo_postal=data['codigo_postal'], 
                    email=data['email'], 
                    contacto_emergencia=data['contacto_emergencia'], 
                    emergency_phone=data['emergency_phone'], 
                    sangre=data['sangre'], 
                    pay_day=data['pay_day'], 
                    horario=data['horario'], 
                    suburbio=data['suburbio'], 
                    direccion=data['direccion'], 
                    condicion_medica=data['condicion_medica'],
                    plan_inscripcion=plan_inscripcion,
                    status=data['status'],
                    sucursal=sucursal_a_usar
                )
                messages.success(request, 'Cliente actualizado con éxito.')
            else:
                # Crear nuevo cliente
                save_client = Clientes(
                    nombre=data['nombre'], 
                    genero=genero, 
                    apellido_paterno=data['apellido_paterno'], 
                    apellido_materno=data['apellido_materno'], 
                    celular=data['celular'], 
                    fecha_nacimiento=data['fecha_nacimiento'],
                    codigo_postal=data['codigo_postal'], 
                    email=data['email'], 
                    contacto_emergencia=data['contacto_emergencia'], 
                    emergency_phone=data['emergency_phone'], 
                    sangre=data['sangre'], 
                    pay_day=data['pay_day'], 
                    horario=data['horario'], 
                    suburbio=data['suburbio'], 
                    direccion=data['direccion'], 
                    condicion_medica=data['condicion_medica'],
                    plan_inscripcion=plan_inscripcion,
                    status=data['status'],
                    sucursal=sucursal_a_usar
                )
                save_client.save()
                messages.success(request, 'Cliente creado con éxito.')

            resp['status'] = 'success'
                
        except Exception as e:
            resp['msg'] = str(e)
            resp['status'] = 'failed'

    return JsonResponse(resp)



@login_required
def manage_products(request):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    product = None
    # Filtrar categorías que están activas y pertenecen a la sucursal del usuario
    categories = Category.objects.filter(status=1, sucursal=sucursal_a_usar)

    if request.method == 'GET':
        data = request.GET
        product_id = data.get('id', '')
        if product_id.isnumeric():
            # Asegurarse de que el producto pertenece a la sucursal del perfil del usuario
            product = Products.objects.filter(id=product_id, sucursal=sucursal_a_usar).first()
            if not product:
                return render(request, 'posApp/error.html', {
                    'message': 'Producto no encontrado o no pertenece a su sucursal.'
                })

    context = {
        'product': product,
        'categories': categories
    }

    return render(request, 'posApp/manage_product.html', context)


@login_required
def save_product(request):
    data = request.POST
    resp = {'status': 'failed'}

    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        resp['msg'] = 'Error: el usuario no tiene un perfil asignado.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        resp['msg'] = 'Error: el perfil no tiene una sucursal asignada.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            resp['msg'] = 'Sucursal no válida en la sesión.'
            return HttpResponse(json.dumps(resp), content_type="application/json")

    product_id = data.get('id', '')
    product_code = data.get('code', '')

    # Verificar si existe un código de producto duplicado dentro de la misma sucursal
    existing_products_query = Products.objects.filter(code=product_code, sucursal=sucursal_a_usar)
    if product_id.isnumeric():
        existing_products_query = existing_products_query.exclude(id=product_id)

    if existing_products_query.exists():
        resp['msg'] = "El código de producto ya existe en la base de datos."
        return HttpResponse(json.dumps(resp), content_type="application/json")

    category_id = data.get('category_id')
    category = Category.objects.filter(id=category_id, sucursal=sucursal_a_usar).first()
    if not category:
        resp['msg'] = "Categoría no encontrada o no pertenece a tu sucursal."
        return HttpResponse(json.dumps(resp), content_type="application/json")

    try:
        product_attributes = {
            'code': product_code,
            'category_id': category,
            'name': data.get('name', ''),
            'description': data.get('description', ''),
            'price': float(data.get('price', '0')),
            'status': int(data.get('status', '1')),
            'sucursal': sucursal_a_usar
        }
        if product_id.isnumeric() and int(product_id) > 0:
            # Actualizar producto existente
            Products.objects.filter(id=product_id, sucursal=sucursal_a_usar).update(**product_attributes)
        else:
            # Crear un nuevo producto
            product = Products(**product_attributes)
            product.save()
        resp['status'] = 'success'
        messages.success(request, 'Producto guardado con éxito.')
    except Exception as e:
        resp['status'] = 'failed'
        resp['msg'] = str(e)

    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def delete_product(request):
    data = request.POST
    resp = {'status': ''}

    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        resp['msg'] = 'Error: el usuario no tiene un perfil asignado.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        resp['msg'] = 'Error: el perfil no tiene una sucursal asignada.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            resp['msg'] = 'Sucursal no válida en la sesión.'
            resp['status'] = 'failed'
            return HttpResponse(json.dumps(resp), content_type="application/json")

    product_id = data.get('id', '')
    if not product_id.isnumeric():
        resp['msg'] = 'ID inválido.'
        resp['status'] = 'failed'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    try:
        # Solo permite eliminar el producto si pertenece a la sucursal del usuario
        product = Products.objects.filter(id=product_id, sucursal=sucursal_a_usar).first()
        if product:
            product.delete()
            resp['status'] = 'success'
            messages.success(request, 'Producto eliminado con éxito.')
        else:
            resp['status'] = 'failed'
            resp['msg'] = 'Producto no encontrado o no pertenece a su sucursal.'
    except Exception as e:
        resp['status'] = 'failed'
        resp['msg'] = f"Error al eliminar el producto: {str(e)}"

    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def pos(request):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    # Filtrar productos y clientes que están activos y pertenecen a la sucursal del usuario
    products = Products.objects.filter(status=1, sucursal=sucursal_a_usar)
    clients = Clientes.objects.filter(status=1, sucursal=sucursal_a_usar)
    tipoPago = FormaPago.objects.all()  # Asumiendo que FormaPago no es específico de sucursal

    # Convertir los productos y clientes en JSON para usar en la plantilla
    product_json = [{'id': product.id, 'name': product.name, 'price': float(product.price)} for product in products]
    client_json = [{'id': client.id} for client in clients]

    context = {
        'page_title': "Punto de Ventas",
        'products': products,
        'clients': clients,
        'product_json': json.dumps(product_json),
        'client_json': json.dumps(client_json),
        'tipo_pago': tipoPago
    }

    return render(request, 'posApp/pos.html', context)

@login_required
def checkout_modal(request):
   
    grand_total = 0
    if 'grand_total' in request.GET:
        grand_total = request.GET['grand_total']
    context = {
        'grand_total' : grand_total,
       
    }
    return render(request, 'posApp/checkout.html',context)

@login_required
def save_pos(request):
    resp = {'status': 'failed', 'msg': ''}
    data = request.POST

    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        resp['msg'] = 'Error: el usuario no tiene un perfil asignado.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        resp['msg'] = 'Error: el perfil no tiene una sucursal asignada.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            resp['msg'] = 'Sucursal no válida en la sesión.'
            return HttpResponse(json.dumps(resp), content_type="application/json")

    # Generar código único usando 'T', la fecha y la hora actual
    now = datetime.now()
    code = 'T' + now.strftime('%Y%m%d%H%M%S')

    # Verificar que el cliente y el tipo de pago existan en la sucursal
    client = Clientes.objects.filter(id=data['client'], sucursal=sucursal_a_usar).first()
    tipoPago = FormaPago.objects.filter(id=data['tipoPago']).first()

    if not client or not tipoPago:
        resp['msg'] = 'Cliente o tipo de pago no encontrado o no pertenece a su sucursal.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    try:
        # Guardar el objeto Sales con el nuevo código
        sales = Sales.objects.create(
            code=code,
            sub_total=data['sub_total'],
            tax=data['tax'],
            tax_amount=data['tax_amount'],
            grand_total=data['grand_total'],
            tendered_amount=data['tendered_amount'],
            tendered_amount_card=data['tendered_amount_card'],
            client=client,
            amount_change=data['amount_change'],
            tipoPago=tipoPago,
            usuario=request.user,  # Usamos el usuario actual
            comentario=data['comentario'],
            sucursal=sucursal_a_usar  # Asegurarse de asignar la sucursal
        )
        sale_id = sales.pk

        # Crear items de venta
        for i, product_id in enumerate(data.getlist('product_id[]')):
            product = Products.objects.filter(id=product_id, sucursal=sucursal_a_usar).first()
            if not product:
                continue  # Ignorar productos que no pertenecen a la sucursal del usuario
            qty = data.getlist('qty[]')[i]
            price = data.getlist('price[]')[i]
            total = float(qty) * float(price)
            salesItems.objects.create(
                sale_id=sales,
                client=client,
                product_id=product,
                qty=qty,
                price=price,
                total=total
            )

        resp['status'] = 'success'
        resp['sale_id'] = sale_id
        messages.success(request, "La venta ha sido guardada.")
    except Exception as e:
        resp['msg'] = "Ocurrió un error: " + str(e)
        print("Unexpected error:", sys.exc_info()[0])

    return HttpResponse(json.dumps(resp), content_type="application/json")


from django.utils import timezone
from datetime import timedelta
import pytz
from decimal import Decimal
from django.db.models import Sum

@login_required
def salesList(request):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Obtener la zona horaria de Ciudad de México
    tz = pytz.timezone('America/Mexico_City')

    # Obtener la fecha y hora actual en la zona horaria de Ciudad de México
    now = timezone.now().astimezone(tz)
    today = now.date()
    tomorrow = today + timedelta(days=1)

    # Definir las fechas de inicio y fin usando la zona horaria correcta
    start_datetime = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.min.time()), tz)
    end_datetime = timezone.make_aware(timezone.datetime.combine(tomorrow, timezone.datetime.max.time()), tz)

    # Asegurarte de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    # Filtrar todos los datos según la sucursal del usuario
    clientes = Clientes.objects.filter(sucursal=sucursal_a_usar)
    formapago = FormaPago.objects.all()  # Asumiendo que FormaPago no es específico de sucursal
    productos = Products.objects.filter(sucursal=sucursal_a_usar)
    categorias = Category.objects.filter(sucursal=sucursal_a_usar)
    tipos_inscripcion = PlanInscripcion.objects.filter(sucursal=sucursal_a_usar)

    sales_query = Sales.objects.filter(sucursal=sucursal_a_usar)
    salidas_query = Salida.objects.filter(sucursal=sucursal_a_usar)

    # Definir las fechas de inicio y fin, y convertirlas a objetos datetime

    start_date = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', tomorrow.strftime('%Y-%m-%d'))

    # Convertir las fechas al formato correcto para las consultas
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()

    # Filtrar ventas y salidas por rango de fechas
    sales_query = sales_query.filter(date_added__date__range=[start_date, end_date])
    salidas_query = salidas_query.filter(fecha__range=[start_date, end_date])

    # Cálculos de totales
    totalEfectivo = Decimal(sales_query.filter(date_added__date=today, tipoPago=1).aggregate(Sum('grand_total'))['grand_total__sum'] or 0)
    totalBanco = Decimal(sales_query.filter(date_added__date=today, tipoPago=2).aggregate(Sum('grand_total'))['grand_total__sum'] or 0)
    totalSalidas = Decimal(salidas_query.aggregate(Sum('monto'))['monto__sum'] or 0)

    tipo_pago_dict = {
        1: "Efectivo",
        2: "Banco",
        3: "Efectivo y Tarjeta"
    }

    # Filtros adicionales basados en parámetros de consulta
    filters = {
        'cliente_id': 'client_id',
        'formapago_id': 'tipoPago',
        'producto_id': 'id__in',
        'categoria_id': 'id__in',
        'tipo_inscripcion_id': 'id__in'
    }
    for param, filter_key in filters.items():
        value = request.GET.get(param)
        if value:
            if param == 'producto_id':
                sales_query = sales_query.filter(**{filter_key: salesItems.objects.filter(product_id=value, product_id__sucursal=sucursal_a_usar).values_list('sale_id', flat=True)})
            elif param == 'categoria_id':
                sales_query = sales_query.filter(**{filter_key: salesItems.objects.filter(product_id__category_id=value, product_id__sucursal=sucursal_a_usar).values_list('sale_id', flat=True)})
            elif param == 'tipo_inscripcion_id':
                sales_query = sales_query.filter(**{filter_key: salesItems.objects.filter(product_name__icontains=PlanInscripcion.objects.get(id=value, sucursal=sucursal_a_usar).nombre).values_list('sale_id', flat=True)})
            else:
                sales_query = sales_query.filter(**{filter_key: value})

    total_money = Decimal(sales_query.aggregate(Sum('grand_total'))['grand_total__sum'] or 0)
    totalSalidas = Decimal(salidas_query.aggregate(Sum('monto'))['monto__sum'] or 0)

    sales = []
    for sale in sales_query:
        client = Clientes.objects.get(id=sale.client_id) if sale.client_id else None
        sales.append({
            'id': sale.id,
            'date': sale.date_added.astimezone(tz).strftime('%Y-%m-%d %H:%M'),  # Convertir la fecha a la zona horaria de Ciudad de México
            'client_name': f"{client.nombre } {client.apellido_paterno} {client.apellido_materno}" if client else 'No Client',
            'horario': client.horario if client else 'No Schedule',
            'plan_inscripcion': client.plan_inscripcion,
            'total': sale.grand_total,
            'payment_type': tipo_pago_dict.get(sale.tipoPago_id, "Unknown") if sale.tipoPago else "Unknown"
        })

    # Preparar el contexto para el template
    context = {
        'page_title': 'Sales Transactions',
        'sales_data': sales,
        'salida_data': list(salidas_query),
        'total_ventas': total_money,
        'total_salidas': totalSalidas,
        'neto': total_money - totalSalidas,
        'totalEfectivo': totalEfectivo,
        'totalBanco': totalBanco,
        'clientes': clientes,
        'formapago': formapago,
        'productos': productos,
        'categorias': categorias,
        'tipos_inscripcion': tipos_inscripcion,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
    }

    return render(request, 'posApp/sales.html', context)

from django.views.decorators.http import require_GET
@require_GET
def buscar_clientes(request):
    query = request.GET.get('query', '')

    if query:
        clientes = Clientes.objects.filter(nombre__icontains=query)[:10]  # Limitar a los primeros 10 resultados
        clientes_data = [{
            'id': cliente.id,
            'nombre': cliente.nombre,
            'apellido_materno': cliente.apellido_materno,
            'apellido_paterno': cliente.apellido_paterno
        } for cliente in clientes]

        return JsonResponse({'clientes': clientes_data})
    
    return JsonResponse({'clientes': []})


@login_required
def cliente_mensualidades(request, cliente_id):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    # Asegurarse de que el cliente pertenezca a la sucursal del usuario
    cliente = get_object_or_404(Clientes, id=cliente_id, sucursal=sucursal_a_usar)
    
    # Filtrar mensualidades por cliente asegurando que se considere la sucursal
    mensualidades = Mensualidad.objects.filter(cliente=cliente).order_by('-fecha_vencimiento')

    context = {
        'cliente': cliente,
        'mensualidades': mensualidades,
    }
    
    return render(request, 'posApp/cliente_mensualidades.html', context)

@login_required
def eliminar_mensualidad(request, cliente_id, mensualidad_id):
    mensualidad = get_object_or_404(Mensualidad, id=mensualidad_id, cliente_id=cliente_id)
    if request.method == 'POST':
        mensualidad.delete()
        messages.success(request, 'Mensualidad eliminada exitosamente.')
    return redirect('cliente_mensualidades', cliente_id=cliente_id)

@login_required
def receipt(request):
    sale_id = request.GET.get('id')

    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    # Asegúrate de que la venta pertenezca a la sucursal del usuario
    sales = get_object_or_404(Sales, id=sale_id, sucursal=sucursal_a_usar)
    
    # Extraer todos los campos de la transacción
    transaction = {}
    for field in Sales._meta.get_fields():
        if field.related_model is None:
            transaction[field.name] = getattr(sales, field.name)
    if 'tax_amount' in transaction:
        transaction['tax_amount'] = format(float(transaction['tax_amount']))

    # Asegurarse de que los items de la venta también pertenecen a la misma venta y sucursal
    ItemList = salesItems.objects.filter(sale_id=sales)

    context = {
        "transaction": transaction,
        "salesItems": ItemList,
        "sales": sales
    }

    return render(request, 'posApp/receipt.html', context)

    # return HttpResponse('')
@login_required
@user_passes_test(lambda u: u.userprofile.is_manager)
def delete_sale(request):
    resp = {'status': 'failed', 'msg': ''}
    sale_id = request.POST.get('id')

    if not sale_id or not sale_id.isnumeric():
        resp['msg'] = "Invalid sale ID."
        return HttpResponse(json.dumps(resp), content_type='application/json')

    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        resp['msg'] = 'Error: el usuario no tiene un perfil asignado.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Asegurarse de que el perfil tiene una sucursal asignada
    if not user_profile.sucursal:
        resp['msg'] = 'Error: el perfil no tiene una sucursal asignada.'
        return HttpResponse(json.dumps(resp), content_type="application/json")

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            resp['msg'] = 'Sucursal no válida en la sesión.'
            return HttpResponse(json.dumps(resp), content_type="application/json")

    try:
        # Verificar que la venta pertenece a la sucursal del usuario antes de intentar eliminarla
        sale = Sales.objects.filter(id=sale_id, sucursal=sucursal_a_usar).first()
        if sale:
            sale.delete()
            resp['status'] = 'success'
            messages.success(request, 'La venta ha sido eliminada.')
        else:
            resp['msg'] = "Venta no encontrada o no pertenece a su sucursal."
            resp['status'] = 'failed'
    except Exception as e:
        resp['msg'] = f"An error occurred: {str(e)}"
        print("Unexpected error:", sys.exc_info()[0])

    return HttpResponse(json.dumps(resp), content_type='application/json')

@login_required
def add_salida(request):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    if request.method == 'POST':
        form = SalidaForm(request.POST)
        if form.is_valid():
            salida = form.save(commit=False)
            # Asignar la sucursal del perfil del usuario a la salida antes de guardar
            salida.sucursal = sucursal_a_usar
            salida.save()
            return redirect('add_salida')
    else:
        form = SalidaForm()

    return render(request, 'posApp/add_salida.html', {'form': form})


@login_required
@transaction.atomic
def pagar_primera_mensualidad(request, cliente_id):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    # Asegúrate de que el cliente pertenece a la sucursal del usuario
    cliente = get_object_or_404(Clientes, id=cliente_id, sucursal=sucursal_a_usar)

    form = MensualidadForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        mensualidad = form.save(commit=False)
        mensualidad.cliente = cliente
        mensualidad.pagado = True
        mensualidad.fecha_pago = timezone.now()
        mensualidad.sucursal = cliente.sucursal
        mensualidad.save()

        tendered_amount_cash = form.cleaned_data.get('tendered_amount_cash', 0)
        tendered_amount_card = form.cleaned_data.get('tendered_amount_card', 0)
        total_paid = float(tendered_amount_cash + tendered_amount_card)
        monto_mensualidad = float(mensualidad.monto)
        cambio = total_paid - monto_mensualidad

        # Obtener la fecha de registro si el usuario es gerente
        fecha_registro = request.POST.get('fecha_registro')
        if fecha_registro and user_profile.is_manager:
            date_added = datetime.datetime.strptime(fecha_registro, '%Y-%m-%d')
        else:
            date_added = timezone.now()

        plan_inscripcion_id = request.POST.get('plan_inscripcion')
        plan_inscripcion = get_object_or_404(PlanInscripcion, id=plan_inscripcion_id, sucursal=sucursal_a_usar)

        # Registrar el pago en el sistema de ventas
        venta = Sales.objects.create(
            code=f"M-{mensualidad.id}-{timezone.now().strftime('%Y%m%d')}",
            sub_total=monto_mensualidad,
            grand_total=monto_mensualidad,
            tax_amount=0,
            sucursal=cliente.sucursal,
            tax=0,
            tendered_amount=tendered_amount_cash,
            tendered_amount_card=tendered_amount_card,
            amount_change=cambio,
            date_added=date_added,  # Usar la fecha proporcionada o la actual
            client=cliente,
            usuario=request.user,
            tipoPago=get_object_or_404(FormaPago, id=request.POST.get('forma_pago')),
            comentario=f"Pago de mensualidad {plan_inscripcion.nombre}"
        )

        # Registrar un SalesItem para el plan de inscripción
        salesItems.objects.create(
            sale_id=venta,
            product_name=plan_inscripcion.nombre,
            price=plan_inscripcion.precio,
            qty=1,
            total=plan_inscripcion.precio,
            client=cliente,
            sucursal=cliente.sucursal
        )

        messages.success(request, 'Mensualidad registrada exitosamente.')
        return redirect('sales-page')

    formas_pago = FormaPago.objects.all()
    planes_inscripcion = PlanInscripcion.objects.filter(sucursal=sucursal_a_usar)

    return render(request, 'posApp/pagar_primera_mensualidad.html', {
        'form': form,
        'cliente': cliente,
        'formas_pago': formas_pago,
        'planes_inscripcion': planes_inscripcion,
    })


@login_required
def seleccionar_mensualidades(request, cliente_id):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    year = datetime.now().year

    # Asegúrate de que el cliente pertenece a la sucursal del usuario
    cliente = get_object_or_404(Clientes, id=cliente_id, sucursal=sucursal_a_usar)

    # Obtener las formas de pago y planes de inscripción para la sucursal del usuario
    formas_pago = FormaPago.objects.all()
    planes_de_inscripcion = PlanInscripcion.objects.filter(sucursal=sucursal_a_usar)

    # Generar una lista de meses disponibles
    meses_disponibles = [f"{year}-{str(m).zfill(2)}" for m in range(1, 13)]

    return render(request, 'posApp/multiples_mensualidades.html', {
        'cliente': cliente,
        'planes_inscripcion': planes_de_inscripcion,
        'formas_pago': formas_pago,
        'meses_disponibles': meses_disponibles,
    })

@login_required
@transaction.atomic
def generar_mensualidades(request, cliente_id):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, 'posApp/error.html', {'message': 'Este usuario no tiene un perfil asociado.'})

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        return render(request, 'posApp/error.html', {'message': 'Este perfil no tiene una sucursal asignada.'})

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return render(request, 'posApp/error.html', {'message': 'Sucursal no válida en la sesión.'})

    # Obtener el cliente y validar que pertenece a la sucursal correcta
    cliente = get_object_or_404(Clientes, id=cliente_id, sucursal=sucursal_a_usar)

    if request.method == 'POST':
        plan_inscripcion_id = request.POST.get('plan_inscripcion')
        meses_seleccionados = request.POST.getlist('meses[]')
        dia_cobro = request.POST.get('dia')
        tendered_amount_cash = float(request.POST.get('tendered_amount_cash', 0))
        tendered_amount_card = float(request.POST.get('tendered_amount_card', 0))

        plan_inscripcion = get_object_or_404(PlanInscripcion, id=plan_inscripcion_id, sucursal=sucursal_a_usar)
        total_pagado = tendered_amount_cash + tendered_amount_card
        total_monto = len(meses_seleccionados) * float(plan_inscripcion.precio)
        cambio = total_pagado - total_monto

        # Crear una única venta para todas las mensualidades seleccionadas
        venta = Sales.objects.create(
            code=f"MULTI-{cliente.id}-{timezone.now().strftime('%Y%m%d')}",
            sub_total=total_monto,
            grand_total=total_monto,
            tendered_amount=tendered_amount_cash,
            tendered_amount_card=tendered_amount_card,
            amount_change=cambio,
            client=cliente,
            sucursal=sucursal_a_usar,
            usuario=request.user,
            tipoPago=get_object_or_404(FormaPago, id=request.POST.get('forma_pago')),
            comentario=f"Pago de múltiples mensualidades {plan_inscripcion.nombre}",
        )

        # Para cada mes seleccionado, crear la mensualidad y añadirla como SalesItem
        for mes in meses_seleccionados:
            # Definir la fecha de vencimiento
            fecha_vencimiento = datetime.strptime(f"{mes}-{dia_cobro}", "%Y-%m-%d")

            # Crear la nueva Mensualidad
            mensualidad = Mensualidad.objects.create(
                cliente=cliente,
                pagado=True,
                fecha_vencimiento=fecha_vencimiento,
                monto=plan_inscripcion.precio,
                sucursal=sucursal_a_usar,
                fecha_pago=timezone.now(),
                plan_inscripcion=plan_inscripcion
            )

            # Añadir cada mensualidad como un SalesItem en la venta
            salesItems.objects.create(
                sale_id=venta,
                product_name=f"Mensualidad {plan_inscripcion.nombre} ({mes})",
                price=plan_inscripcion.precio,
                qty=1,
                total=plan_inscripcion.precio,
                client=cliente,
                sucursal=sucursal_a_usar
            )

        messages.success(request, "Mensualidades generadas y registradas en una sola venta exitosamente.")
        return redirect('sales-page')

    return redirect('error-page')

@login_required
def create_plan_inscripcion(request):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        messages.error(request, 'No se puede crear el plan, usuario sin perfil asociado.')
        return redirect('error_page')  # Redirige a una página de error o maneja el error apropiadamente

    # Asegúrate de que el perfil tenga una sucursal asignada
    if not user_profile.sucursal:
        messages.error(request, 'No se puede crear el plan, usuario sin sucursal asignada.')
        return redirect('error_page')

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            messages.error(request, 'Sucursal no válida en la sesión.')
            return redirect('error_page')

    if request.method == 'POST':
        form = PlanInscripcionForm(request.POST)
        if form.is_valid():
            plan = form.save(commit=False)
            # Asegurarse de que el plan esté asociado a la sucursal determinada
            plan.sucursal = sucursal_a_usar
            plan.save()
            messages.success(request, f'Plan "{plan.nombre}" creado exitosamente.')
            return redirect('create_plan_inscripcion')
    else:
        form = PlanInscripcionForm()

    return render(request, 'posApp/create_plan_inscripcion.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_superuser or u.userprofile.is_manager)
def create_sucursal(request):
    if request.method == 'POST':
        form = SucursalForm(request.POST)
        if form.is_valid():
            sucursal = form.save()
            messages.success(request, f'Sucursal "{sucursal.nombre}" creada exitosamente.')
            return redirect('create_sucursal')
    else:
        form = SucursalForm()

    return render(request, 'posApp/create_sucursal.html', {'form': form})

from django.utils.dateparse import parse_date
from openpyxl import Workbook
@login_required
def export_sales_to_excel(request):
    try:
        # Intentamos obtener el perfil del usuario
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return HttpResponse("Error: el usuario no tiene un perfil asignado.", status=403)

    # Asegúrate de que el perfil tiene una sucursal asignada
    if not user_profile.sucursal:
        return HttpResponse("Error: el usuario no tiene una sucursal asignada.", status=403)

    # Determinar la sucursal a usar
    sucursal_a_usar = user_profile.sucursal
    if user_profile.is_manager and 'sucursal_id' in request.session:
        try:
            sucursal_a_usar = Sucursal.objects.get(id=request.session['sucursal_id'])
        except Sucursal.DoesNotExist:
            return HttpResponse("Error: la sucursal seleccionada no es válida.", status=400)

    today = timezone.now().date()
    start_date_str = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('end_date', today.strftime('%Y-%m-%d'))

    # Convertir las fechas de cadena a objetos date
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)

    if start_date is None or end_date is None:
        return HttpResponse("Error: formato de fecha inválido. Use 'YYYY-MM-DD'.", status=400)

    # Filtrar las ventas y salidas por la sucursal del usuario
    sales_query = Sales.objects.filter(date_added__range=[start_date, end_date], sucursal=sucursal_a_usar)
    salidas_query = Salida.objects.filter(fecha__range=[start_date, end_date], sucursal=sucursal_a_usar)

    # Crear el libro y la hoja de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Añadir encabezados
    headers = ['#', 'Fecha', 'Detalle', 'Cliente/Concepto', 'Realizado Por', 'Total']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)

    # Añadir datos de ventas
    for row_num, sale in enumerate(sales_query, 2):
        ws.cell(row=row_num, column=1, value=sale.id)
        ws.cell(row=row_num, column=2, value=sale.date_added.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_num, column=3, value='Venta')
        ws.cell(row=row_num, column=4, value=f"{sale.client.nombre} {sale.client.apellido_paterno} {sale.client.apellido_materno}")
        ws.cell(row=row_num, column=5, value=sale.usuario.username)
        ws.cell(row=row_num, column=6, value=sale.grand_total)

    # Añadir datos de salidas
    for index, salida in enumerate(salidas_query, start=len(list(sales_query)) + 2):
        ws.cell(row=index, column=1, value=salida.id)
        ws.cell(row=index, column=2, value=salida.fecha.strftime('%Y-%m-%d'))
        ws.cell(row=index, column=3, value='Salida')
        ws.cell(row=index, column=4, value=salida.concepto)
        ws.cell(row=index, column=5, value='N/A')
        ws.cell(row=index, column=6, value=-salida.monto)

    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)

    return response
